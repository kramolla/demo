from connection import Base, engine, session
from models import *

from openpyxl import load_workbook

BASE_DIR = "excel/"
product_types_file = BASE_DIR + "Product_type_import.xlsx"
material_types_file = BASE_DIR + "Material_type_import.xlsx"
products_file = BASE_DIR + "Products_import.xlsx"
partners_file = BASE_DIR + "Partners_import.xlsx"
orders_file = BASE_DIR + "Partner_products_import.xlsx"

Base.metadata.drop_all(engine)
Base.metadata.create_all(engine)

try:
    wb = load_workbook(product_types_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, coef = row

        if not name or not coef:
            continue

        session.add(ProductTypeModel(
            name=name,
            coef=coef
        ))
        session.commit()

    print("Типы продукции - OK")

    wb = load_workbook(material_types_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, coef = row

        if not name or not coef:
            continue

        session.add(MaterialTypeModel(
            name=name,
            coef=coef
        ))
        session.commit()

    print("Типы материалов - OK")

    wb = load_workbook(products_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_type_name, name, article, min_cost = row

        if not product_type_name or not name or not article or not min_cost:
            continue

        product_type = session.query(ProductTypeModel).filter_by(name=product_type_name).first()

        session.add(ProductModel(
            name=name,
            article=article,
            min_cost=min_cost,
            product_type=product_type.id
        ))
        session.commit()

    print("Продукты - OK")

    wb = load_workbook(partners_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        type, name, boss_name, email, phone_number, address, inn, rate = row

        if not type or not name or not boss_name or not email or not phone_number\
                or not address or not inn or not rate:
            continue



        session.add(PartnerModel(
            type=type,
            name=name,
            boss_name=boss_name,
            email=email,
            phone_number=phone_number,
            address=address,
            inn=inn,rate=rate
        ))
        session.commit()

    print("Партнёры - OK")

    wb = load_workbook(orders_file)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        product_name, partner_name, count, date = row

        if not product_name or not partner_name or not count or not date:
            continue

        product = session.query(ProductModel).filter_by(name=product_name).first()
        partner = session.query(PartnerModel).filter_by(name=partner_name).first()

        session.add(OrderModel(
            product=product.id,
            partner=partner.id,
            count=count,
            date=date
        ))
        session.commit()

    print("Партнёры - OK")

except Exception as e:
    session.rollback()
    print(f"Ты дебил: {e}")
finally:
    session.close()
    print("Импорт завершён")